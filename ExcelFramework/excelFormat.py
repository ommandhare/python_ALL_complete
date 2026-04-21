from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Create a workbook and select the active data_ws
wb = Workbook()
ws = wb.active
ws.title = "Formatted Sheet"
#ws.sheet_view.showGridLines = False
# Add some sample data
ws['A1'] = "Name"
ws['B1'] = "Age"
ws['A2'] = "Alice"
ws['B2'] = 30
ws['B10'] = 100



# Format header (row 1)
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4F81BD")  # Blue background
alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = alignment
    cell.border = thin_border

# Format data rows
for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=2):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left")

# Auto-adjust column width
for column_cells in ws.columns:
    max_length = 0
    column = column_cells[0].column_letter
    for cell in column_cells:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[column].width = max_length + 2

# Save workbook
wb.save("formatted_excel.xlsx")


## commendted code
"""
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

# Create a workbook and data_ws
wb = Workbook()
ws = wb.active

# Sample data
ws.append(["Name", "Score"])
ws.append(["Alice", 45])
ws.append(["Bob", 78])
ws.append(["Charlie", 60])
ws.append(["David", 30])

# Define fill for highlighting
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Apply conditional formatting to cells in B2:B100 if greater than 50
ws.conditional_formatting.add(
    "B2:B100",
    CellIsRule(operator='greaterThan', formula=['50'], fill=red_fill)
)

# Save the file
wb.save("conditional_formatting.xlsx")

"""

# commented code snippet 2
"""
from openpyxl.formatting.rule import FormulaRule

# Highlight rows where score < 40
ws.conditional_formatting.add(
    "A2:B100",
    FormulaRule(formula=["$B2<40"], fill=red_fill)
)

"""

# commented code snippet 3
"""
from openpyxl.formatting.rule import ColorScaleRule

color_scale_rule = ColorScaleRule(
    start_type='min', start_color='FFFFFF',
    mid_type='percentile', mid_value=50, mid_color='FFFF00',
    end_type='max', end_color='FF0000'
)

ws.conditional_formatting.add("B2:B100", color_scale_rule)


"""

## commented code snippet 4

"""
from openpyxl import Workbook

# Create a workbook and select active sheet
wb = Workbook()
ws = wb.active

# Rename the sheet (optional)
ws.title = "No Gridlines"

# Add some sample data
ws['A1'] = "Name"
ws['B1'] = "Age"
ws['A2'] = "Alice"
ws['B2'] = 30

# Turn off gridlines
ws.sheet_view.showGridLines = False

# Save the workbook
wb.save("no_gridlines.xlsx")


"""