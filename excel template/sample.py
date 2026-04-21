from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ------------------------------
# Helper styles
# ------------------------------
header_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
pass_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
bold_font = Font(bold=True)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(border_style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)


# ------------------------------
# Create workbook
# ------------------------------
wb = Workbook()

# -----------------------------------------------
# SHEET 1: SUMMARY TEMPLATE (Your first screenshot)
# -----------------------------------------------
ws1 = wb.active
ws1.title = "Summary"

ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 50

rows = [
    ("Test Case Name", "UTSC_Shipper_cost_updated"),
    ("Test Case Description", "This test case shows the unit testing of Shipper_cost_updated"),
    ("sheet 1 ","Contains Primary information"),
    ("sheet 2 ","Join info of temp_shipper"),
    ("sheet 3 ","Source/Target temp_whs_latest_avbl_data"),
    ("sheet 4 ","Source/Target temp_item_latest_avbl_data")
]

for r, (k, v) in enumerate(rows, start=2):
    ws1[f"A{r}"].value = k
    ws1[f"B{r}"].value = v
    ws1[f"A{r}"].font = bold_font
    ws1[f"A{r}"].alignment = left
    ws1[f"B{r}"].alignment = left
    ws1[f"A{r}"].border = border
    ws1[f"B{r}"].border = border

#
# # -----------------------------------------------
# # SHEET 2: PRIMARY TEST CASE TABLE (Your second screenshot)
# # -----------------------------------------------
ws2 = wb.create_sheet("Primary")

cols = ["Test Case ID", "Test Case Description", "Steps to Execute",
        "Expected Result", "Actual Result", "Status (Pass/Fail)", "Remarks"]

ws2.append(cols)

# Styling header
for col in range(1, len(cols)+1):
    cell = ws2.cell(row=1, column=col)
    cell.font = bold_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
    ws2.column_dimensions[get_column_letter(col)].width = 25

# Sample rows (blank structure)
data_rows = [
    ["TC_001", "Validate SQL syntax",
     "1. Write SQL query.\n2. Execute query.",
     "Query executes without syntax errors.", "", "Pass", "N/A"],

    ["TC_002", "Validate joins and relationships",
     "Validate data between related tables.",
     "Correct relationship mapping.", "", "Pass", "Correct joins."],

    ["TC_005", "Test data filtering",
     "Apply WHERE conditions.\nExecute query.",
     "Query filters correctly.", "", "Pass", "Adjust for partial matches."]
]

for row in data_rows:
    ws2.append(row)

# Apply alignment + borders
for row in ws2.iter_rows(min_row=2, max_col=7):
    for cell in row:
        cell.border = border
        cell.alignment = left
        if cell.column == 6:     # STATUS column
            cell.fill = pass_fill
            cell.alignment = center


# # -----------------------------------------------
# # SHEET 3: TABLE ROW COUNT (third screenshot)
# # -----------------------------------------------
ws3 = wb.create_sheet("Row_Count")

ws3["A1"] = "TABLE ROW COUNT"
ws3.merge_cells("A1:B1")
ws3["A1"].font = Font(bold=True, size=14)
ws3["A1"].alignment = center

ws3.append(["Table Query", "Rows Count"])
ws3["A2"].font = bold_font
ws3["B2"].font = bold_font

ws3["A3"] = "`gcp-abs-sdim-dev-prj-01.sdim_ds_data_analytics_dev_cts.shipper_master`"
ws3["B3"] = 87845

ws3["A3"].alignment = left
ws3["B3"].alignment = center

# Border formatting
for row in ws3.iter_rows(min_row=2, max_col=2, max_row=3):
    for cell in row:
        cell.border = border


# -----------------------------------------------
# SAVE FILE
# -----------------------------------------------
wb.save(r"C:\Users\Om Mandhare\PycharmProjects\python_ALL_complete\excel template\Test_Template.xlsx")
print("Excel Template Created Successfully: Test_Template.xlsx")
