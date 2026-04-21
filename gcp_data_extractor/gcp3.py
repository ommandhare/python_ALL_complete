from google.cloud import bigquery
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

ROOT_DIR = "./root_dir"
OUTPUT_FOLDER = "./excel_with_python_test"

if not os.path.isdir(OUTPUT_FOLDER):
    os.mkdir(OUTPUT_FOLDER)

client = bigquery.Client.from_service_account_json("./data-cleaning-428505-51dbd39732d7.json")

def apply_excel_formatting(output_file):
    wb = load_workbook(output_file)
    ws = wb.active

    # Styles
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Format all other cells with border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border

    # Auto column width
    for column in ws.columns:
        max_length = 0
        col = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col].width = max_length + 2

    # ---- NEW PART: Freeze header if rows > 21 ----
    if ws.max_row > 21:
        ws.freeze_panes = "A2"

    wb.save(output_file)


def getResults(file):
    try:
        with open(file, 'r') as f:
            query = f.read()
        print(f"\nRunning query from {file}")
        df = client.query(query).to_dataframe()

        # CAPITALIZE HEADERS
        df.columns = [c.upper() for c in df.columns]

        name = os.path.splitext(os.path.basename(file))[0]
        output_file = os.path.join(OUTPUT_FOLDER, f"{name}.xlsx")
        df.to_excel(output_file, index=False, engine='openpyxl')

        # Apply formatting
        apply_excel_formatting(output_file)

        print(f"Results saved to {output_file}")

    except Exception as e:
        print(f"Something went wrong: {e}")



for root, dirs, files in os.walk(ROOT_DIR):
    for file in files:
        full_path = os.path.join(root, file)
        getResults(full_path)
